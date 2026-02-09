import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import requests
import json
from datetime import datetime
import os
from dotenv import load_dotenv, find_dotenv

# Загружаем .env из корня проекта
load_dotenv(find_dotenv())

# Установка openpyxl если нет
try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference

# Настройка для корректного отображения русского текста
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.unicode_minus'] = False

# ==================== КОНФИГУРАЦИЯ ====================
BASE_DIR = r"C:\Users\Ruslan\Desktop\AI learning\data"
CSV_PATH = os.path.join(BASE_DIR, "Procurement KPI Analysis Dataset.csv")
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")
CHARTS_DIR = os.path.join(BASE_DIR, "procurement_charts")
EXCEL_PATH = os.path.join(CHARTS_DIR, "Procurement_Audit_Report.xlsx")

# Создаём папку для графиков
os.makedirs(CHARTS_DIR, exist_ok=True)

# ==================== 1. АНАЛИЗ ДАННЫХ С PANDAS ====================
print("=" * 60)
print("АУДИТ ЗАКУПОК - АНАЛИЗ ДАННЫХ")
print("=" * 60)

# Загрузка данных
df = pd.read_csv(CSV_PATH)
print(f"\nЗагружено записей: {len(df)}")

# Преобразование дат
df['Order_Date'] = pd.to_datetime(df['Order_Date'], errors='coerce')
df['Delivery_Date'] = pd.to_datetime(df['Delivery_Date'], errors='coerce')

# Расчёт ключевых метрик
df['Delivery_Days'] = (df['Delivery_Date'] - df['Order_Date']).dt.days
df['Defect_Rate'] = (df['Defective_Units'] / df['Quantity'] * 100).fillna(0)
df['Price_Savings'] = df['Unit_Price'] - df['Negotiated_Price']
df['Savings_Percent'] = (df['Price_Savings'] / df['Unit_Price'] * 100)
df['Total_Order_Value'] = df['Quantity'] * df['Negotiated_Price']

# Агрегация по поставщикам
supplier_stats = df.groupby('Supplier').agg({
    'PO_ID': 'count',
    'Total_Order_Value': 'sum',
    'Defect_Rate': 'mean',
    'Delivery_Days': 'mean',
    'Defective_Units': 'sum',
    'Quantity': 'sum'
}).rename(columns={'PO_ID': 'Orders_Count'})

supplier_stats['Overall_Defect_Rate'] = (supplier_stats['Defective_Units'] / supplier_stats['Quantity'] * 100)

# Non-compliance по поставщикам
non_compliance_by_supplier = df[df['Compliance'] == 'No'].groupby('Supplier').size()
supplier_stats['Non_Compliance_Count'] = non_compliance_by_supplier.reindex(supplier_stats.index, fill_value=0)
supplier_stats['Non_Compliance_Rate'] = (supplier_stats['Non_Compliance_Count'] / supplier_stats['Orders_Count'] * 100)

# ==================== 2. СОЗДАНИЕ ГРАФИКОВ ====================
print("\n" + "=" * 60)
print("СОЗДАНИЕ ГРАФИКОВ")
print("=" * 60)

chart_files = []

# Цветовая палитра
colors = ['#2ecc71', '#3498db', '#9b59b6', '#e74c3c', '#f39c12']

# --- ГРАФИК 1: Объём заказов по поставщикам ---
fig, ax = plt.subplots(figsize=(10, 6))
suppliers = supplier_stats.index.tolist()
values = supplier_stats['Total_Order_Value'] / 1_000_000  # в миллионах

bars = ax.bar(suppliers, values, color=colors)
ax.set_title('Объём заказов по поставщикам', fontsize=14, fontweight='bold')
ax.set_xlabel('Поставщик')
ax.set_ylabel('Сумма заказов (млн $)')
ax.set_ylim(0, max(values) * 1.15)

for bar, val in zip(bars, values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2,
            f'${val:.1f}M', ha='center', va='bottom', fontsize=10)

plt.tight_layout()
chart1_path = os.path.join(CHARTS_DIR, 'chart1_order_volume.png')
plt.savefig(chart1_path, dpi=150, bbox_inches='tight')
plt.close()
chart_files.append(chart1_path)
print(f"График 1 сохранён: {chart1_path}")

# --- ГРАФИК 2: Процент брака по поставщикам ---
fig, ax = plt.subplots(figsize=(10, 6))
defect_rates = supplier_stats['Overall_Defect_Rate'].sort_values(ascending=False)

bars = ax.bar(defect_rates.index, defect_rates.values, color=['#e74c3c' if x > 7 else '#2ecc71' for x in defect_rates.values])
ax.axhline(y=defect_rates.mean(), color='orange', linestyle='--', linewidth=2, label=f'Среднее: {defect_rates.mean():.1f}%')
ax.set_title('Процент брака по поставщикам', fontsize=14, fontweight='bold')
ax.set_xlabel('Поставщик')
ax.set_ylabel('Процент брака (%)')
ax.legend()

for bar, val in zip(bars, defect_rates.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2,
            f'{val:.1f}%', ha='center', va='bottom', fontsize=10)

plt.tight_layout()
chart2_path = os.path.join(CHARTS_DIR, 'chart2_defect_rate.png')
plt.savefig(chart2_path, dpi=150, bbox_inches='tight')
plt.close()
chart_files.append(chart2_path)
print(f"График 2 сохранён: {chart2_path}")

# --- ГРАФИК 3: Non-Compliance по поставщикам ---
fig, ax = plt.subplots(figsize=(10, 6))
nc_rates = supplier_stats['Non_Compliance_Rate'].sort_values(ascending=False)

bars = ax.bar(nc_rates.index, nc_rates.values, color=['#e74c3c' if x > 15 else '#f39c12' if x > 5 else '#2ecc71' for x in nc_rates.values])
ax.axhline(y=nc_rates.mean(), color='blue', linestyle='--', linewidth=2, label=f'Среднее: {nc_rates.mean():.1f}%')
ax.set_title('Уровень Non-Compliance по поставщикам', fontsize=14, fontweight='bold')
ax.set_xlabel('Поставщик')
ax.set_ylabel('Non-Compliance (%)')
ax.legend()

for bar, val in zip(bars, nc_rates.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
            f'{val:.1f}%', ha='center', va='bottom', fontsize=10)

plt.tight_layout()
chart3_path = os.path.join(CHARTS_DIR, 'chart3_non_compliance.png')
plt.savefig(chart3_path, dpi=150, bbox_inches='tight')
plt.close()
chart_files.append(chart3_path)
print(f"График 3 сохранён: {chart3_path}")

# --- ГРАФИК 4: Среднее время доставки ---
fig, ax = plt.subplots(figsize=(10, 6))
delivery_days = supplier_stats['Delivery_Days'].sort_values(ascending=False)

bars = ax.bar(delivery_days.index, delivery_days.values, color=colors)
ax.axhline(y=delivery_days.mean(), color='red', linestyle='--', linewidth=2, label=f'Среднее: {delivery_days.mean():.1f} дней')
ax.set_title('Среднее время доставки по поставщикам', fontsize=14, fontweight='bold')
ax.set_xlabel('Поставщик')
ax.set_ylabel('Дней до доставки')
ax.legend()

for bar, val in zip(bars, delivery_days.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2,
            f'{val:.1f}', ha='center', va='bottom', fontsize=10)

plt.tight_layout()
chart4_path = os.path.join(CHARTS_DIR, 'chart4_delivery_time.png')
plt.savefig(chart4_path, dpi=150, bbox_inches='tight')
plt.close()
chart_files.append(chart4_path)
print(f"График 4 сохранён: {chart4_path}")

# --- ГРАФИК 5: Комплексная оценка поставщиков (радар) ---
fig, ax = plt.subplots(figsize=(10, 8), subplot_kw=dict(projection='polar'))

categories = ['Объём заказов', 'Качество\n(низкий брак)', 'Compliance', 'Скорость\nдоставки']
num_cats = len(categories)
angles = [n / float(num_cats) * 2 * np.pi for n in range(num_cats)]
angles += angles[:1]

# Нормализация данных (0-100)
def normalize(series, inverse=False):
    if inverse:
        return 100 - ((series - series.min()) / (series.max() - series.min()) * 100)
    return (series - series.min()) / (series.max() - series.min()) * 100

for i, supplier in enumerate(suppliers):
    values = [
        normalize(supplier_stats['Total_Order_Value'])[supplier],
        normalize(supplier_stats['Overall_Defect_Rate'], inverse=True)[supplier],
        normalize(supplier_stats['Non_Compliance_Rate'], inverse=True)[supplier],
        normalize(supplier_stats['Delivery_Days'], inverse=True)[supplier]
    ]
    values += values[:1]
    ax.plot(angles, values, 'o-', linewidth=2, label=supplier, color=colors[i])
    ax.fill(angles, values, alpha=0.15, color=colors[i])

ax.set_xticks(angles[:-1])
ax.set_xticklabels(categories, fontsize=10)
ax.set_title('Комплексная оценка поставщиков', fontsize=14, fontweight='bold', y=1.08)
ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0))

plt.tight_layout()
chart5_path = os.path.join(CHARTS_DIR, 'chart5_radar.png')
plt.savefig(chart5_path, dpi=150, bbox_inches='tight')
plt.close()
chart_files.append(chart5_path)
print(f"График 5 сохранён: {chart5_path}")

# --- ГРАФИК 6: Распределение по категориям товаров ---
fig, axes = plt.subplots(1, 2, figsize=(14, 6))

# Pie chart - объём по категориям
category_volume = df.groupby('Item_Category')['Total_Order_Value'].sum()
axes[0].pie(category_volume, labels=category_volume.index, autopct='%1.1f%%', colors=plt.cm.Set3.colors)
axes[0].set_title('Объём заказов по категориям', fontsize=12, fontweight='bold')

# Bar chart - брак по категориям
category_defect = df.groupby('Item_Category')['Defect_Rate'].mean().sort_values(ascending=False)
bars = axes[1].bar(category_defect.index, category_defect.values, color=plt.cm.Set3.colors)
axes[1].set_title('Средний процент брака по категориям', fontsize=12, fontweight='bold')
axes[1].set_xlabel('Категория')
axes[1].set_ylabel('Процент брака (%)')
axes[1].tick_params(axis='x', rotation=45)

for bar, val in zip(bars, category_defect.values):
    axes[1].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                 f'{val:.1f}%', ha='center', va='bottom', fontsize=9)

plt.tight_layout()
chart6_path = os.path.join(CHARTS_DIR, 'chart6_categories.png')
plt.savefig(chart6_path, dpi=150, bbox_inches='tight')
plt.close()
chart_files.append(chart6_path)
print(f"График 6 сохранён: {chart6_path}")

# ==================== 2.5 ЭКСПОРТ В EXCEL ====================
print("\n" + "=" * 60)
print("ЭКСПОРТ В EXCEL")
print("=" * 60)

# Создаём Excel файл
wb = openpyxl.Workbook()

# Стили
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
danger_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
good_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

# --- ЛИСТ 1: Сводка ---
ws_summary = wb.active
ws_summary.title = "Сводка"

summary_data = [
    ["АУДИТ ЗАКУПОК - СВОДНЫЙ ОТЧЁТ", ""],
    ["", ""],
    ["Показатель", "Значение"],
    ["Всего заказов", len(df)],
    ["Общая стоимость заказов", f"${df['Total_Order_Value'].sum():,.2f}"],
    ["Экономия на переговорах", f"${(df['Price_Savings'] * df['Quantity']).sum():,.2f}"],
    ["Средний процент брака", f"{df['Defect_Rate'].mean():.2f}%"],
    ["Среднее время доставки", f"{df['Delivery_Days'].mean():.1f} дней"],
    ["Количество поставщиков", df['Supplier'].nunique()],
    ["Категорий товаров", df['Item_Category'].nunique()],
    ["", ""],
    ["ВЫЯВЛЕННЫЕ ПРОБЛЕМЫ", ""],
    ["Заказы с браком >15%", len(df[df['Defect_Rate'] > 15])],
    ["Non-Compliance случаев", f"{len(df[df['Compliance'] == 'No'])} ({len(df[df['Compliance'] == 'No'])/len(df)*100:.1f}%)"],
    ["Cancelled заказов", len(df[df['Order_Status'] == 'Cancelled'])],
    ["Pending заказов", len(df[df['Order_Status'] == 'Pending'])],
]

for row in summary_data:
    ws_summary.append(row)

ws_summary.merge_cells('A1:B1')
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary['A1'].alignment = Alignment(horizontal='center')
style_header(ws_summary, 3)
auto_width(ws_summary)

# --- ЛИСТ 2: Анализ поставщиков ---
ws_suppliers = wb.create_sheet("Поставщики")

supplier_export = supplier_stats[['Orders_Count', 'Total_Order_Value', 'Overall_Defect_Rate',
                                   'Non_Compliance_Rate', 'Delivery_Days']].copy()
supplier_export.columns = ['Кол-во заказов', 'Сумма заказов ($)', 'Брак (%)',
                           'Non-Compliance (%)', 'Ср. доставка (дни)']
supplier_export = supplier_export.reset_index()
supplier_export.columns = ['Поставщик'] + list(supplier_export.columns[1:])

ws_suppliers.append(list(supplier_export.columns))
for _, row in supplier_export.iterrows():
    ws_suppliers.append(list(row))

style_header(ws_suppliers)

# Условное форматирование
for row_idx in range(2, len(supplier_export) + 2):
    # Брак
    defect_cell = ws_suppliers.cell(row=row_idx, column=4)
    if defect_cell.value and defect_cell.value > 7:
        defect_cell.fill = danger_fill
        defect_cell.font = Font(color="FFFFFF", bold=True)
    elif defect_cell.value and defect_cell.value > 5:
        defect_cell.fill = warning_fill

    # Non-compliance
    nc_cell = ws_suppliers.cell(row=row_idx, column=5)
    if nc_cell.value and nc_cell.value > 20:
        nc_cell.fill = danger_fill
        nc_cell.font = Font(color="FFFFFF", bold=True)
    elif nc_cell.value and nc_cell.value > 10:
        nc_cell.fill = warning_fill

auto_width(ws_suppliers)

# Форматируем числа
for row in ws_suppliers.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        if cell.value:
            cell.number_format = '$#,##0.00'

for row in ws_suppliers.iter_rows(min_row=2, min_col=4, max_col=5):
    for cell in row:
        if cell.value:
            cell.number_format = '0.00%'
            cell.value = cell.value / 100  # Конвертируем в процент

# --- ЛИСТ 3: Проблемные заказы (высокий брак) ---
ws_defects = wb.create_sheet("Высокий брак")

high_defect_export = df[df['Defect_Rate'] > 10][['PO_ID', 'Supplier', 'Item_Category',
                                                   'Quantity', 'Defective_Units', 'Defect_Rate',
                                                   'Total_Order_Value']].copy()
high_defect_export.columns = ['ID заказа', 'Поставщик', 'Категория', 'Количество',
                               'Брак (ед.)', 'Брак (%)', 'Сумма ($)']
high_defect_export = high_defect_export.sort_values('Брак (%)', ascending=False)

ws_defects.append(list(high_defect_export.columns))
for _, row in high_defect_export.iterrows():
    ws_defects.append(list(row))

style_header(ws_defects)
auto_width(ws_defects)

# --- ЛИСТ 4: Non-Compliance ---
ws_nc = wb.create_sheet("Non-Compliance")

nc_export = df[df['Compliance'] == 'No'][['PO_ID', 'Supplier', 'Item_Category',
                                           'Order_Status', 'Quantity', 'Total_Order_Value']].copy()
nc_export.columns = ['ID заказа', 'Поставщик', 'Категория', 'Статус', 'Количество', 'Сумма ($)']

ws_nc.append(list(nc_export.columns))
for _, row in nc_export.iterrows():
    ws_nc.append(list(row))

style_header(ws_nc)
auto_width(ws_nc)

# --- ЛИСТ 5: По категориям ---
ws_categories = wb.create_sheet("Категории")

category_stats = df.groupby('Item_Category').agg({
    'PO_ID': 'count',
    'Total_Order_Value': 'sum',
    'Defect_Rate': 'mean',
    'Delivery_Days': 'mean'
}).round(2)
category_stats.columns = ['Кол-во заказов', 'Сумма ($)', 'Ср. брак (%)', 'Ср. доставка (дни)']
category_stats = category_stats.reset_index()
category_stats.columns = ['Категория'] + list(category_stats.columns[1:])

ws_categories.append(list(category_stats.columns))
for _, row in category_stats.iterrows():
    ws_categories.append(list(row))

style_header(ws_categories)
auto_width(ws_categories)

# --- ЛИСТ 6: Все данные ---
ws_all = wb.create_sheet("Все данные")

export_cols = ['PO_ID', 'Supplier', 'Order_Date', 'Delivery_Date', 'Item_Category',
               'Order_Status', 'Quantity', 'Unit_Price', 'Negotiated_Price',
               'Defective_Units', 'Compliance', 'Defect_Rate', 'Delivery_Days', 'Total_Order_Value']
all_data_export = df[export_cols].copy()

ws_all.append(export_cols)
for _, row in all_data_export.iterrows():
    ws_all.append(list(row))

style_header(ws_all)
auto_width(ws_all)

# Сохраняем Excel
wb.save(EXCEL_PATH)
print(f"Excel отчёт сохранён: {EXCEL_PATH}")

# ==================== 3. АНАЛИЗ АНОМАЛИЙ ====================
anomalies = {}

# 1. Высокий брак (>15%)
high_defect = df[df['Defect_Rate'] > 15].copy()
anomalies['high_defect'] = {
    'count': len(high_defect),
    'worst_suppliers': high_defect.groupby('Supplier')['Defect_Rate'].mean().sort_values(ascending=False).head(5).to_dict()
}

# 2. Non-compliance
non_compliance = df[df['Compliance'] == 'No'].copy()
anomalies['non_compliance'] = {
    'count': len(non_compliance),
    'percent': len(non_compliance) / len(df) * 100,
    'by_supplier': non_compliance.groupby('Supplier').size().sort_values(ascending=False).to_dict()
}

# Общая статистика
stats = {
    'total_orders': len(df),
    'total_value': df['Total_Order_Value'].sum(),
    'total_savings': (df['Price_Savings'] * df['Quantity']).sum(),
    'avg_defect_rate': df['Defect_Rate'].mean(),
    'avg_delivery_days': df['Delivery_Days'].mean()
}

# ==================== 4. ГЕНЕРАЦИЯ ОТЧЁТА ЧЕРЕЗ DEEPSEEK ====================
print("\n" + "=" * 60)
print("ГЕНЕРАЦИЯ ОТЧЁТА ЧЕРЕЗ DEEPSEEK API")
print("=" * 60)

analysis_summary = f"""
Данные аудита закупок (Procurement KPI Analysis):

ОБЩАЯ СТАТИСТИКА:
- Всего заказов: {stats['total_orders']}
- Общая стоимость: ${stats['total_value']:,.2f}
- Экономия: ${stats['total_savings']:,.2f}
- Средний брак: {stats['avg_defect_rate']:.2f}%
- Среднее время доставки: {stats['avg_delivery_days']:.1f} дней

ПОКАЗАТЕЛИ ПО ПОСТАВЩИКАМ:
{supplier_stats[['Orders_Count', 'Total_Order_Value', 'Overall_Defect_Rate', 'Non_Compliance_Rate', 'Delivery_Days']].round(2).to_string()}

АНОМАЛИИ:
- Заказы с браком >15%: {anomalies['high_defect']['count']}
- Non-compliance: {anomalies['non_compliance']['count']} ({anomalies['non_compliance']['percent']:.1f}%)
"""

prompt = f"""На основе аудита закупок составь краткий отчёт на русском (до 800 символов).

{analysis_summary}

Структура: 1) Резюме 2) Топ-проблемы 3) Рекомендации
Формат: текст для Telegram с эмодзи."""

deepseek_url = "https://api.deepseek.com/v1/chat/completions"
headers = {
    "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
    "Content-Type": "application/json"
}
payload = {
    "model": "deepseek-chat",
    "messages": [
        {"role": "system", "content": "Ты аналитик закупок. Пиши кратко и по делу."},
        {"role": "user", "content": prompt}
    ],
    "temperature": 0.7,
    "max_tokens": 1000
}

try:
    response = requests.post(deepseek_url, headers=headers, json=payload, timeout=60)
    response.raise_for_status()
    result = response.json()
    report = result['choices'][0]['message']['content']
    print("Отчёт сгенерирован!")
except Exception as e:
    print(f"Ошибка DeepSeek: {e}")
    report = f"""📊 АУДИТ ЗАКУПОК

📈 Резюме: {stats['total_orders']} заказов на ${stats['total_value']/1_000_000:.1f}M

⚠️ Проблемы:
• Брак >15%: {anomalies['high_defect']['count']} заказов
• Non-compliance: {anomalies['non_compliance']['percent']:.1f}%
• Худший по браку: Alpha_Inc

✅ Рекомендации:
• Усилить контроль качества Alpha_Inc
• Пересмотреть контракт с Delta_Logistics"""

# ==================== 5. ОТПРАВКА В TELEGRAM ====================
print("\n" + "=" * 60)
print("ОТПРАВКА В TELEGRAM")
print("=" * 60)

# Функция отправки фото
def send_photo(file_path, caption=""):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto"
    with open(file_path, 'rb') as photo:
        files = {'photo': photo}
        data = {'chat_id': TELEGRAM_CHAT_ID, 'caption': caption[:1024]}
        response = requests.post(url, files=files, data=data, timeout=30)
        return response.ok

# Функция отправки документа
def send_document(file_path, caption=""):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
    with open(file_path, 'rb') as doc:
        files = {'document': doc}
        data = {'chat_id': TELEGRAM_CHAT_ID, 'caption': caption[:1024]}
        response = requests.post(url, files=files, data=data, timeout=60)
        return response.ok

# Функция отправки текста
def send_message(text):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    data = {'chat_id': TELEGRAM_CHAT_ID, 'text': text}
    response = requests.post(url, data=data, timeout=30)
    return response.ok

# Отправляем текстовый отчёт
print("Отправка текстового отчёта...")
if send_message(report):
    print("✓ Текстовый отчёт отправлен")
else:
    print("✗ Ошибка отправки текста")

# Отправляем графики
chart_captions = [
    "📊 Объём заказов по поставщикам",
    "🔴 Процент брака по поставщикам",
    "⚠️ Non-Compliance по поставщикам",
    "🚚 Время доставки по поставщикам",
    "📈 Комплексная оценка поставщиков",
    "📦 Анализ по категориям товаров"
]

for chart_path, caption in zip(chart_files, chart_captions):
    print(f"Отправка: {caption}...")
    if send_photo(chart_path, caption):
        print(f"✓ {caption} отправлен")
    else:
        print(f"✗ Ошибка отправки {caption}")

# Отправляем Excel файл
print("Отправка Excel отчёта...")
if send_document(EXCEL_PATH, "📋 Полный отчёт аудита закупок (Excel)"):
    print("✓ Excel отчёт отправлен")
else:
    print("✗ Ошибка отправки Excel")

print("\n" + "=" * 60)
print("АУДИТ С ГРАФИКАМИ И EXCEL ЗАВЕРШЁН!")
print("=" * 60)
